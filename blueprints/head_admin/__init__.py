from flask import Blueprint, render_template, request, flash, redirect, url_for, session
from models import TicketTypeType, TicketType, Admin, User, Ticket
from functools import wraps
import datetime
from database import db_session
import openpyxl

h_admin_bp = Blueprint("h_admin_bp", __name__, template_folder="templates", static_folder="static")


def check_head_admin(func):
    @wraps(func)
    def inner(*args, **kwargs):
        if "permit" in session:
            if session["permit"] == 1:
                return func(*args, **kwargs)
        return redirect(url_for("admin_bp.check_tickets_view"))
    return inner


@h_admin_bp.route("/operations/")
@check_head_admin
def operations():
    ttts = TicketTypeType.query.all()
    return render_template("head_admin/head_operations.html", ttts=ttts)


@h_admin_bp.route("/add-piece/", methods=['POST'])
@check_head_admin
def add_piece():
    name = request.form["piece-name"]
    speaker = request.form["piece-speaker"]
    start = datetime.datetime.strptime(request.form["piece-start"], '%H:%M').time()
    max_cap = int(request.form["piece-cap"])
    ttt = TicketTypeType.query.filter(TicketTypeType.id == int(request.form["piece-ttt"])).first()

    tt = TicketType(name, speaker, max_cap)
    tt.start = start
    tt.ticket_type_type = ttt

    db_session.add(tt)
    db_session.commit()

    flash(f"Nová časť programu v {ttt.name} bola pridaná", "success")
    return redirect(url_for("h_admin_bp.stats"))


@h_admin_bp.route("/edit-piece/", methods=['POST'])
@check_head_admin
def edit_piece():
    name = request.form["piece-name"]
    speaker = request.form["piece-speaker"]
    start = request.form["piece-start"]
    max_cap = int(request.form["piece-cap"])
    tt = TicketType.query.filter(TicketType.id == int(request.form["piece-id"])).first()

    tt.name = name
    tt.speaker = speaker
    if len(start)>5:
        start = datetime.datetime.strptime(start, '%H:%M:%S').time()
    else:
        start = datetime.datetime.strptime(start, '%H:%M').time()
    tt.start = start
    tt.max_cap = max_cap

    db_session.commit()

    flash(f"Časť programu '{tt.name}' bola upravená", "success")
    return redirect(url_for("h_admin_bp.stats"))


@h_admin_bp.route("/delete-piece/<int:piece_id>/")
@check_head_admin
def delete_piece(piece_id):
    tt = TicketType.query.filter(TicketType.id == piece_id).first()

    db_session.delete(tt)
    db_session.commit()

    flash(f"Časť programu '{tt.name}' bola vymazaná", "success")
    return redirect(url_for("h_admin_bp.stats"))


@h_admin_bp.route('add-admin', methods=['POST'])
@check_head_admin
def add_admin():
    admin = Admin(request.form["admin-name"], request.form["admin-email"], 2)
    admin.set_password(request.form["admin-password"])

    db_session.add(admin)
    db_session.commit()

    flash("Nový admin bol pridaný", "success")
    return redirect(url_for("h_admin_bp.stats"))


@h_admin_bp.route("/stats/")
@check_head_admin
def stats():
    non_confirm = len(User.query.filter(User.confirm == False).all())
    all_users = len(User.query.all())
    arrived_users = len(User.query.filter(User.active_places).all())
    ttts = TicketTypeType.query.filter()

    return render_template("head_admin/stats.html", non_confirm=non_confirm, all_users=all_users,
                           arrived_users=arrived_users, ttts=ttts)


@h_admin_bp.route("/stats/<int:piece_id>/")
@check_head_admin
def stats_piece(piece_id):
    tt = TicketType.query.filter(TicketType.id == piece_id).first()
    confirmed_users = "broken"

    return render_template("head_admin/stats_piece.html", tt=tt, confirmed_users=confirmed_users)


@h_admin_bp.route("/all-users/")
@check_head_admin
def all_users():
    users = User.query.all()
    return render_template("head_admin/users_list.html", users=users)


@h_admin_bp.route("/confirm-users/")
@check_head_admin
def confirm_users():
    users = User.query.filter(User.confirm == True).all()
    return render_template("head_admin/users_list.html", users=users)


@h_admin_bp.route("/non-confirm-users/")
@check_head_admin
def non_confirm_users():
    users = User.query.filter(User.confirm == False).all()
    return render_template("head_admin/users_list.html", users=users)


@h_admin_bp.route("/arrived-users/")
@check_head_admin
def arrived_users():
    users = User.query.filter(User.active_places).all()
    return render_template("head_admin/users_list.html", users=users)


@h_admin_bp.route("/add-users-from-excel/", methods=['POST'])
@check_head_admin
def add_users_excel():
    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file:
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active

        try:
            start_row = int(request.form["start-row"])
            col_name = int(request.form["col-name"])
            col_email = int(request.form["col-email"])
            col_age = int(request.form["col-age"])
            col_place = int(request.form["col-place"])
            col_who = int(request.form["col-who"])
            col_where = int(request.form["col-where"])
            col_reg = int(request.form["col-reg"])
            col_otp = int(request.form["col-otp"])

            for i in range(start_row - 1, sheet_obj.max_row):
                name = sheet_obj.cell(row=i + 1, column=col_name).value
                email = sheet_obj.cell(row=i + 1, column=col_email).value
                age = sheet_obj.cell(row=i + 1, column=col_age).value
                place = sheet_obj.cell(row=i + 1, column=col_place).value
                who = sheet_obj.cell(row=i + 1, column=col_who).value
                where = sheet_obj.cell(row=i + 1, column=col_where).value
                reg = sheet_obj.cell(row=i + 1, column=col_reg).value
                otp = sheet_obj.cell(row=i + 1, column=col_otp).value

                user = User(name, email)
                user.age = age
                user.city = place
                user.otp = otp
                user.who = who
                user.where = where
                user.confirm = False

                db_session.add(user)

                tt = TicketType.query.filter(TicketType.name == "TEDA #trnavská_veda").first()
                ticket = Ticket(tt, user)
                db_session.add(ticket)

                if len(reg) > 30:
                    tt2 = TicketType.query.filter(TicketType.name == "Diskuttujme o vzdelávaní").first()
                    ticket2 = Ticket(tt2, user)
                    db_session.add(ticket2)

                db_session.commit()

        except:
            flash("Nastala chyba pri nahrávaní. Ujistite sa, či dáta z tabuľky nie su už v systéme", "danger")
            return redirect(url_for("h_admin_bp.stats"))

        flash("Uživatelia z execelu boli úspešne pridané", "success")
        return redirect(url_for("h_admin_bp.stats"))

    return redirect(url_for("h_admin_bp.stats"))


@h_admin_bp.route("/add-users-from-excel-wor/", methods=['POST'])
@check_head_admin
def add_users_excel_wor():
    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file:
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active


        start_row = int(request.form["start-row-w"])
        col_name = int(request.form["col-name-w"])
        col_email = int(request.form["col-email-w"])
        col_age = int(request.form["col-age-w"])
        col_place = int(request.form["col-place-w"])
        col_who = int(request.form["col-who-w"])
        col_wor_1 = int(request.form["col-w-1"])
        col_wor_2 = int(request.form["col-w-2"])
        col_wor_3 = int(request.form["col-w-3"])
        col_otp = int(request.form["col-otp-w"])
        col_where = int(request.form["col-where-w"])
        for i in range(start_row - 1, sheet_obj.max_row):
            email = sheet_obj.cell(row=i + 1, column=col_email).value
            user = User.query.filter(User.email == email).first()
            if not user:
                name = sheet_obj.cell(row=i + 1, column=col_name).value
                age = sheet_obj.cell(row=i + 1, column=col_age).value
                place = sheet_obj.cell(row=i + 1, column=col_place).value
                who = sheet_obj.cell(row=i + 1, column=col_who).value
                where = sheet_obj.cell(row=i + 1, column=col_where).value
                otp = sheet_obj.cell(row=i + 1, column=col_otp).value
                user = User(name, email)
                user.age = age
                user.city = place
                user.otp = otp
                user.who = who
                user.where = where
                user.confirm = False
                db_session.add(user)
                db_session.commit()
            wor1 = sheet_obj.cell(row=i + 1, column=col_wor_1).value
            wor2 = sheet_obj.cell(row=i + 1, column=col_wor_2).value
            wor3 = sheet_obj.cell(row=i + 1, column=col_wor_3).value

            tt1 = TicketType.query.filter(TicketType.name == wor1).first()
            if tt1:
                ticket = Ticket(tt1, user)
                db_session.add(ticket)
            tt2 = TicketType.query.filter(TicketType.name == wor2).first()
            if tt2:
                ticket = Ticket(tt2, user)
                db_session.add(ticket)
            tt3 = TicketType.query.filter(TicketType.name == wor3).first()
            if tt3:
                ticket = Ticket(tt3, user)
                db_session.add(ticket)
            db_session.commit()

        flash("Uživatelia z execelu boli úspešne pridané", "success")
        return redirect(url_for("h_admin_bp.stats"))

    return redirect(url_for("h_admin_bp.stats"))

